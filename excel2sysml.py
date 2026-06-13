"""要件定義ExcelファイルからSysMLv2テキストを生成
"""
import sys
from pathlib import Path
from openpyxl import load_workbook


def convert_sysmlv2(ws, package_name):
    """ExcelシートからSysMLv2を表示
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print('empty sheet')
        return

    # 一行目はヘッダ
    data_rows = rows[1:]

    # SysMLv2テキストのヘッダ表示
    print(f'''package {package_name} {{

    requirement def RequirementItem {{
        attribute category : String;
    }}

    public import RequirementDerivation::*;

''')

    # 読み込み
    for row in data_rows:
        (r_id, p_ids, r_cate, r_name, r_doc) = row
        print(f"""    requirement <'{r_id}'> '{r_name}' : RequirementItem {{
        category = '{r_cate}';
        doc /* {r_doc} */
    }}""")
        if p_ids:
            for pid in p_ids.split(','):
                print(' '*4 + '#derivation connection {')
                print(" "*8 + "end #original ::> '" + pid + "';")
                print(" "*8 + "end #derive ::> '" + r_id + "';")
                print(' '*4 + '}')

        print()

    print('}')


def main(filepath):
    package_name = filepath.stem
    wb = load_workbook(filepath, data_only=True)
    for name in wb.sheetnames:
        convert_sysmlv2(wb[name], package_name)
        return


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print(f'Usage: python {sys.argv[0]} <excel-file>')
        exit(1)

    main(Path(sys.argv[1]))
